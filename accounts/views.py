# accounts/views.py
# IMPORTANT: This file should ONLY contain views, NO models!

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from .models import Student, DisciplinaryCase, MonthlyReport
from accounts.models import DisciplinaryCase  # Import both models from models.py
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
from django.db.models import Count
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, timedelta
import json
import calendar

def landing_page(request):
    return render(request, 'landing.html')

def admin_login_view(request):
    if request.user.is_authenticated:
        logout(request)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
        
            messages.success(request, f'Welcome back, JDK Supervisor! You have successfully logged in.')
            return redirect('admin_dashboard')
        
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'admin_login.html')

def admin_logout_view(request):
    """Admin logout with notification"""
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        messages.success(request, f'JDK Supervisor has been logged out successfully!')
    else:
        messages.info(request, 'You were not logged in.')
    
    return redirect('landing_page')

def admin_dashboard(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    # Get recent students
    recent_students = Student.objects.all()[:5]
    
    # Get recent disciplinary cases - THIS WAS MISSING FROM CONTEXT
    recent_cases = DisciplinaryCase.objects.select_related('student').order_by('-incident_date')[:5]
    
    total_students = Student.objects.count()
    
    # Get case counts
    pending_cases = DisciplinaryCase.objects.filter(status='pending').count()
    under_review_cases = DisciplinaryCase.objects.filter(status='under_review').count()
    resolved_cases = DisciplinaryCase.objects.filter(status='resolved').count()
    total_cases = DisciplinaryCase.objects.count()

    context = {
        'user': request.user,
        'total_students': total_students,
        'pending_cases': pending_cases,
        'under_review_cases': under_review_cases,
        'resolved_cases': resolved_cases,
        'total_cases': total_cases,
        'recent_students': recent_students,
        'recent_cases': recent_cases,  # ADD THIS LINE - was missing!
    }
    
    return render(request, 'admin_dashboard.html', context)

# Update this in your accounts/views.py

# Replace your student_login_view with this debug version

# Replace your student_login_view in accounts/views.py

def student_login_view(request):
    if request.method == 'POST':
        student_id = request.POST.get('student_id', '').strip()
        ic_number = request.POST.get('ic_number', '').strip()
        
        # Validate input
        if not student_id:
            messages.error(request, 'Please enter your Student ID.')
            return render(request, 'student_login.html')
        
        if not ic_number:
            messages.error(request, 'Please enter your IC Number.')
            return render(request, 'student_login.html')
        
        if len(ic_number) != 12 or not ic_number.isdigit():
            messages.error(request, 'IC Number must be exactly 12 digits.')
            return render(request, 'student_login.html')
        
        try:
            # Try to find the student
            student = Student.objects.get(
                student_id=student_id,
                ic_number=ic_number
            )
            
            # Create or update student login tracking
            try:
                student_login, created = StudentLogin.objects.get_or_create(
                    student=student,
                    defaults={
                        'last_login': timezone.now(),
                        'login_count': 1
                    }
                )
                
                if not created:
                    student_login.last_login = timezone.now()
                    student_login.login_count += 1
                    student_login.save()
            except Exception as login_error:
                # If StudentLogin fails, continue without it
                print(f"StudentLogin error: {login_error}")
            
            # Store student info in session
            request.session['student_id'] = student.id
            request.session['student_name'] = student.name
            request.session['is_student'] = True
            
            messages.success(request, f'Welcome, {student.name}!')
            return redirect('student_profile')
            
        except Student.DoesNotExist:
            messages.error(request, 'Invalid Student ID or IC Number.')
        except Exception as e:
            print(f"Login error: {str(e)}")  # For debugging
            messages.error(request, 'An error occurred during login. Please try again.')
    
    return render(request, 'student_login.html')
# Add this to your accounts/views.py

# Add/replace these views in your accounts/views.py

def student_logout_view(request):
    """Student logout with notification"""
    student_name = request.session.get('student_name', 'Student')
    
    if request.session.get('is_student'):
        # Clear the session
        request.session.flush()
        messages.success(request, f'{student_name} has been logged out successfully!')
    else:
        messages.info(request, 'You were not logged in.')
    
    return redirect('landing_page')


def logout_view(request):
    """General logout view that handles both admin and student"""
    if request.user.is_authenticated:
        # Admin logout
        username = request.user.username
        logout(request)
        messages.success(request, f'JDK Supervisor has been logged out successfully!')
    elif request.session.get('is_student'):
        # Student logout
        student_name = request.session.get('student_name', 'Student')
        request.session.flush()
        messages.success(request, f'{student_name} has been logged out successfully!')
    else:
        messages.info(request, 'You were not logged in.')
    
    return redirect('landing_page')
def add_student_view(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    if request.method == 'POST':
        try:
            # Get form data
            student_name = request.POST.get('student_name')
            student_id = request.POST.get('student_id')
            ic_number = request.POST.get('ic_number')
            phone_number = request.POST.get('phone_number')
            address = request.POST.get('address')
            student_class = request.POST.get('student_class')
            semester = request.POST.get('semester')
            program = request.POST.get('program')
            academic_advisor = request.POST.get('academic_advisor')
            
            # Validate required fields
            if not all([student_name, student_id, ic_number, phone_number, address, 
                       student_class, semester, program, academic_advisor]):
                messages.error(request, 'All fields are required.')
                return render(request, 'add_student.html')
            
            # Check for duplicate Student ID
            if Student.objects.filter(student_id=student_id).exists():
                messages.error(request, f'Student ID "{student_id}" already exists. Please use a different Student ID.')
                return render(request, 'add_student.html')
            
            # Check for duplicate IC Number
            if Student.objects.filter(ic_number=ic_number).exists():
                messages.error(request, f'IC Number "{ic_number}" already exists. Please check the IC Number.')
                return render(request, 'add_student.html')
            
            # Validate IC Number format (12 digits)
            if len(ic_number) != 12 or not ic_number.isdigit():
                messages.error(request, 'IC Number must be exactly 12 digits.')
                return render(request, 'add_student.html')
            
            # Validate phone number format
            if not phone_number.startswith('0') or len(phone_number) < 10 or len(phone_number) > 11 or not phone_number.isdigit():
                messages.error(request, 'Phone number must start with 0 and be 10-11 digits long.')
                return render(request, 'add_student.html')
            
            # Create and save student
            student = Student(
                name=student_name,
                student_id=student_id,
                ic_number=ic_number,
                phone_number=phone_number,
                address=address,
                class_name=student_class,
                semester=semester,
                program=program,
                academic_advisor=academic_advisor
            )
            
            student.save()
            messages.success(request, f'Student {student.name} (ID: {student.student_id}) added successfully!')
            return redirect('add_student')
            
        except Exception as e:
            # Handle any other unexpected errors
            print(f"Unexpected error in add_student_view: {str(e)}")  # For debugging
            messages.error(request, 'An unexpected error occurred. Please try again.')
            return render(request, 'add_student.html')
    
    return render(request, 'add_student.html')

def student_list_view(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    # Get all students
    students = Student.objects.all()
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        students = students.filter(
            Q(name__icontains=search) |
            Q(student_id__icontains=search) |
            Q(ic_number__icontains=search)
        )
    
    context = {
        'students': students,
    }
    
    return render(request, 'student_list.html', context)

def student_details_api(request, student_id):
    """API endpoint for student details modal"""
    student = get_object_or_404(Student, id=student_id)
    data = {
        'name': student.name,
        'student_id': student.student_id,
        'ic_number': student.ic_number,
        'phone_number': student.phone_number,
        'address': student.address,
        'program_display': student.get_program_display(),
        'class_name': student.class_name,
        'semester': student.semester,
        'academic_advisor': student.academic_advisor,
    }
    return JsonResponse(data)

def student_delete_view(request, student_id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        student_name = student.name
        student.delete()
        messages.success(request, f'Student {student_name} has been deleted successfully.')
        return redirect('student_list')
    
    return redirect('student_list')

def student_edit_view(request, student_id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    student = get_object_or_404(Student, id=student_id)
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            new_student_id = request.POST.get('student_id')
            new_ic_number = request.POST.get('ic_number')
            phone_number = request.POST.get('phone_number')
            address = request.POST.get('address')
            program = request.POST.get('program')
            class_name = request.POST.get('class_name')
            semester = request.POST.get('semester')
            academic_advisor = request.POST.get('academic_advisor')
            
            # Validate required fields
            if not all([name, new_student_id, new_ic_number, phone_number, address, 
                       program, class_name, semester, academic_advisor]):
                messages.error(request, 'All fields are required.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Check for duplicate Student ID (excluding current student)
            if Student.objects.filter(student_id=new_student_id).exclude(id=student.id).exists():
                messages.error(request, f'Student ID "{new_student_id}" already exists. Please use a different Student ID.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Check for duplicate IC Number (excluding current student)
            if Student.objects.filter(ic_number=new_ic_number).exclude(id=student.id).exists():
                messages.error(request, f'IC Number "{new_ic_number}" already exists. Please check the IC Number.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Validate IC Number format (12 digits)
            if len(new_ic_number) != 12 or not new_ic_number.isdigit():
                messages.error(request, 'IC Number must be exactly 12 digits.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Validate phone number format
            if not phone_number.startswith('0') or len(phone_number) < 10 or len(phone_number) > 11 or not phone_number.isdigit():
                messages.error(request, 'Phone number must start with 0 and be 10-11 digits long.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Validate semester
            try:
                semester_int = int(semester)
                if semester_int < 1 or semester_int > 6:
                    messages.error(request, 'Semester must be between 1 and 6.')
                    return render(request, 'student_edit.html', {'student': student})
            except ValueError:
                messages.error(request, 'Semester must be a valid number.')
                return render(request, 'student_edit.html', {'student': student})
            
            # Update student data if all validations pass
            student.name = name
            student.student_id = new_student_id
            student.ic_number = new_ic_number
            student.phone_number = phone_number
            student.address = address
            student.program = program
            student.class_name = class_name
            student.semester = semester
            student.academic_advisor = academic_advisor
            
            student.save()
            messages.success(request, f'Student {student.name} (ID: {student.student_id}) updated successfully!')
            return redirect('student_list')
            
        except Exception as e:
            # Handle any other unexpected errors
            print(f"Unexpected error in student_edit_view: {str(e)}")  # For debugging
            messages.error(request, 'An unexpected error occurred. Please try again.')
            return render(request, 'student_edit.html', {'student': student})
    
    return render(request, 'student_edit.html', {'student': student})

# Disciplinary Cases Views

def disciplinary_cases_view(request):
    """Main disciplinary cases page with student profiles and cases"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    # Get all students with their related cases
    students = Student.objects.prefetch_related('disciplinary_cases').all()
    
    # Search functionality
    search = request.GET.get('search')
    if search:
        students = students.filter(
            Q(student_id__icontains=search)
        )
    
    # Filter by case status
    case_status = request.GET.get('case_status')
    if case_status:
        students = students.filter(disciplinary_cases__status=case_status).distinct()
    
    # Filter by case type
    case_type = request.GET.get('case_type')
    if case_type:
        students = students.filter(disciplinary_cases__case_type=case_type).distinct()
    
    context = {
        'students': students,
    }
    
    return render(request, 'disciplinary_case.html', context)

def add_disciplinary_case_view(request):
    """Add new disciplinary case"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    if request.method == 'POST':
        try:
            student_id = request.POST.get('student_id')
            student = get_object_or_404(Student, id=student_id)
            
            # Generate unique case_id
            import uuid
            case_id = f"CASE-{uuid.uuid4().hex[:8].upper()}"
            
            case = DisciplinaryCase(
                case_id=case_id,
                student=student,
                case_type=request.POST.get('case_type'),
                description=request.POST.get('description'),
                incident_date=request.POST.get('incident_date'),
                status=request.POST.get('status', 'pending')
            )
            
            case.save()
            messages.success(request, f'Disciplinary case added for {student.name} successfully!')
            
        except Exception as e:
            messages.error(request, f'Error adding disciplinary case: {str(e)}')
    
    return redirect('disciplinary_cases')

def disciplinary_case_details_api(request, case_id):
    """API endpoint for case details modal"""
    case = get_object_or_404(DisciplinaryCase, id=case_id)
    
    data = {
        'case_type_display': case.get_case_type_display(),
        'status': case.status,
        'status_display': case.get_status_display(),
        'incident_date': case.incident_date.strftime('%B %d, %Y'),
      
        'description': case.description,
        'student_name': case.student.name,
        'student_id': case.student.student_id,
        'student_program': case.student.get_program_display(),
        'student_class': case.student.class_name,
    }
    
    return JsonResponse(data)
def edit_disciplinary_case_view(request, case_id):
    """Edit disciplinary case with mandatory description validation"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    case = get_object_or_404(DisciplinaryCase, id=case_id)
    
    if request.method == 'POST':
        try:
            # Get form data
            case_type = request.POST.get('case_type')
            description = request.POST.get('description', '').strip()  # Strip whitespace
            incident_date = request.POST.get('incident_date')
            status = request.POST.get('status')
            reported_by = request.POST.get('reported_by', '').strip()
            action_taken = request.POST.get('action_taken', '').strip()
            
            # Validate required fields
            if not case_type:
                messages.error(request, 'Case type is required.')
                return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
            
            if not description:
                messages.error(request, 'Description is required and cannot be empty.')
                return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
           
            
            if not incident_date:
                messages.error(request, 'Incident date is required.')
                return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
            
            if not status:
                messages.error(request, 'Status is required.')
                return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
            
            # Validate incident date format
            try:
                from datetime import datetime
                datetime.strptime(incident_date, '%Y-%m-%d')
            except ValueError:
                messages.error(request, 'Invalid incident date format.')
                return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
            
            # Update case fields
            case.case_type = case_type
            case.description = description
            case.incident_date = incident_date
            case.status = status
        
        
            case.save()
            messages.success(request, f'Disciplinary case for {case.student.name} updated successfully!')
            return redirect('disciplinary_cases')
            
        except Exception as e:
            # Handle any other unexpected errors
            print(f"Unexpected error in edit_disciplinary_case_view: {str(e)}")  # For debugging
            messages.error(request, 'An unexpected error occurred. Please try again.')
            return render(request, 'edit_disciplinary_case.html', {'case': case, 'student': case.student})
    
    context = {
        'case': case,
        'student': case.student,
    }
    
    return render(request, 'edit_disciplinary_case.html', context)
def delete_disciplinary_case_view(request, case_id):
    """Delete disciplinary case"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    case = get_object_or_404(DisciplinaryCase, id=case_id)
    
    if request.method == 'POST':
        student_name = case.student.name
        case.delete()
        messages.success(request, f'Disciplinary case for {student_name} has been deleted successfully.')
    
    return redirect('disciplinary_cases')
# views.py - Monthly Report View


@login_required
def monthly_report(request):
    """Monthly report view with filters and statistics"""
    
    # Check if download is requested
    if request.GET.get('format') == 'excel':
        return download_report(request)
    
    # Get filter parameters
    selected_month = request.GET.get('month', str(timezone.now().month))
    selected_year = request.GET.get('year', str(timezone.now().year))
    
    try:
        month = int(selected_month)
        year = int(selected_year)
    except (ValueError, TypeError):
        month = timezone.now().month
        year = timezone.now().year
    
    # Calculate date range for the selected month
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Get cases for the selected month based on incident_date
    cases_queryset = DisciplinaryCase.objects.filter(
        incident_date__gte=start_date,
        incident_date__lt=end_date
    )
    
    # Basic statistics
    total_cases = cases_queryset.count()
    resolved_cases = cases_queryset.filter(status='resolved').count()
    pending_cases = cases_queryset.filter(status='pending').count()
    under_review_cases = cases_queryset.filter(status='under_review').count()
    
    # Recent cases (last 10)
    recent_cases = cases_queryset.select_related('student').order_by('-incident_date')[:10]
    
    # Weekly breakdown
    weekly_data, weekly_labels = get_weekly_breakdown(cases_queryset, start_date, end_date)
    
    # Case type statistics
    case_type_stats = get_case_type_statistics(cases_queryset)
    case_type_labels = [stat['type_display'] for stat in case_type_stats]
    case_type_data = [stat['count'] for stat in case_type_stats]
    
    # Program statistics
    program_stats = get_program_statistics(cases_queryset)
    
    # Month choices for dropdown
    month_choices = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # Year choices for dropdown
    current_year = timezone.now().year
    year_choices = [(y, str(y)) for y in range(current_year - 2, current_year + 3)]
    
    # Generate or update monthly report record
    report, created = MonthlyReport.objects.get_or_create(
        month=f"{calendar.month_name[month]} {year}",
        defaults={
            'report_id': f"RPT-{year}{month:02d}",
            'total_cases': total_cases,
            'active_cases': pending_cases + under_review_cases,
            'resolved_cases': resolved_cases,
            'generated_by': request.user
        }
    )
    
    if not created:
        # Update existing report
        report.total_cases = total_cases
        report.active_cases = pending_cases + under_review_cases
        report.resolved_cases = resolved_cases
        report.save()
    
    context = {
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_name': calendar.month_name[month],
        'month_choices': month_choices,
        'year_choices': year_choices,
        'total_cases': total_cases,
        'resolved_cases': resolved_cases,
        'pending_cases': pending_cases,
        'under_review_cases': under_review_cases,
        'recent_cases': recent_cases,
        'weekly_data': json.dumps(weekly_data),
        'weekly_labels': json.dumps(weekly_labels),
        'case_type_stats': case_type_stats,
        'case_type_labels': json.dumps(case_type_labels),
        'case_type_data': json.dumps(case_type_data),
        'program_stats': program_stats,
        'report': report,
    }
    
    return render(request, 'monthly_report.html', context)
def get_weekly_breakdown(cases_queryset, start_date, end_date):
    """Get weekly breakdown of cases for the month"""
    weekly_data = []
    weekly_labels = []
    
    current_date = start_date
    week_num = 1
    
    while current_date < end_date:
        # Calculate week end date
        week_end = min(current_date + timedelta(days=6), end_date - timedelta(days=1))
        
        # Count cases for this week
        week_cases = cases_queryset.filter(
            incident_date__gte=current_date,
            incident_date__lte=week_end
        ).count()
        
        weekly_data.append(week_cases)
        weekly_labels.append(f"Wk{week_num}")
        
        # Move to next week
        current_date = week_end + timedelta(days=1)
        week_num += 1
    
    return weekly_data, weekly_labels


def get_case_type_statistics(cases_queryset):
    """Get case type statistics with percentages"""
    total_cases = cases_queryset.count()
    
    case_type_stats = []
    case_types = cases_queryset.values('case_type').annotate(
        count=Count('case_type')
    ).order_by('-count')
    
    for case_type in case_types:
        # Get display name for case type
        display_name = dict(DisciplinaryCase.CASE_TYPE_CHOICES).get(
            case_type['case_type'], 
            case_type['case_type']
        )
        
        percentage = round((case_type['count'] / total_cases * 100), 1) if total_cases > 0 else 0
        
        case_type_stats.append({
            'type': case_type['case_type'],
            'type_display': display_name,
            'count': case_type['count'],
            'percentage': percentage
        })
    
    return case_type_stats


def get_program_statistics(cases_queryset):
    """Get program-wise statistics"""
    program_stats = []
    
    # Get unique programs from students involved in cases
    programs = cases_queryset.values('student__program').annotate(
        case_count=Count('id'),
        student_count=Count('student__id', distinct=True)
    ).order_by('-case_count')
    
    for program in programs:
        # Get display name for program
        display_name = dict(Student.PROGRAM_CHOICES).get(
            program['student__program'], 
            program['student__program']
        )
        
        program_stats.append({
            'program': program['student__program'],
            'program_display': display_name,
            'case_count': program['case_count'],
            'student_count': program['student_count']
        })
    
    return program_stats


@login_required
def download_report(request):
    """Download monthly report as Excel"""
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')
    
    month = request.GET.get('month', str(timezone.now().month))
    year = request.GET.get('year', str(timezone.now().year))
    format_type = request.GET.get('format', 'excel')
    
    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        month = timezone.now().month
        year = timezone.now().year
    
    # Calculate date range
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)
    
    # Get cases and statistics
    cases_queryset = DisciplinaryCase.objects.filter(
        incident_date__gte=start_date,
        incident_date__lt=end_date
    ).select_related('student')
    
    if format_type == 'excel':
        return generate_excel_report(request,cases_queryset, month, year)
    else:
        # Fallback to redirect if format not supported
        messages.error(request, 'Unsupported download format.')
        return redirect('monthly_report')


def generate_excel_report(request, cases_queryset, month, year):
    """Generate Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        import io
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Report_{calendar.month_name[month]}_{year}"
        
        # Define styles
        header_font = Font(bold=True, size=16, color="FFFFFF")
        subheader_font = Font(bold=True, size=14)
        table_header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        table_header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        ws['A1'] = f"KPMIM Student Disciplinary System"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Monthly Disciplinary Report - {calendar.month_name[month]} {year}"
        ws['A2'].font = Font(bold=True, size=14)
        ws.merge_cells('A2:H2')
        
        # Statistics Summary
        ws['A4'] = "Summary Statistics"
        ws['A4'].font = subheader_font
        
        # Calculate statistics
        total_cases = cases_queryset.count()
        resolved_cases = cases_queryset.filter(status='resolved').count()
        pending_cases = cases_queryset.filter(status='pending').count()
        under_review_cases = cases_queryset.filter(status='under_review').count()
        
        ws['A5'] = "Total Cases:"
        ws['B5'] = total_cases
        ws['A6'] = "Resolved Cases:"
        ws['B6'] = resolved_cases
        ws['A7'] = "Pending Cases:"
        ws['B7'] = pending_cases
        ws['A8'] = "Under Review:"
        ws['B8'] = under_review_cases
        
        # Case Details Header
        ws['A10'] = "Case Details"
        ws['A10'].font = subheader_font
        
        # Table headers
        headers = ['Case ID', 'Student Name', 'Student ID', 'Program', 'Class', 'Case Type', 'Status', 'Incident Date', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Case Data
        for row, case in enumerate(cases_queryset, 12):
            data = [
                case.case_id,
                case.student.name,
                case.student.student_id,
                case.student.get_program_display(),
                case.student.class_name,
                case.get_case_type_display(),
                case.get_status_display(),
                timezone.localtime(case.incident_date).strftime('%Y-%m-%d'),
                case.description[:150] + '...' if len(case.description) > 150 else case.description
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 8:  # Date column
                    cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        column_widths = {
            'A': 15,  # Case ID
            'B': 25,  # Student Name
            'C': 15,  # Student ID
            'D': 30,  # Program
            'E': 15,  # Class
            'F': 25,  # Case Type
            'G': 15,  # Status
            'H': 20,  # Incident Date
            'I': 50   # Description
        }
        
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width
        
        # Add case type statistics if there are cases
        if total_cases > 0:
            # Case Type Breakdown
            current_row = len(cases_queryset) + 15
            ws[f'A{current_row}'] = "Case Type Breakdown"
            ws[f'A{current_row}'].font = subheader_font
            
            # Headers for case type stats
            current_row += 1
            ws[f'A{current_row}'] = "Case Type"
            ws[f'B{current_row}'] = "Count"
            ws[f'C{current_row}'] = "Percentage"
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{current_row}']
                cell.font = table_header_font
                cell.fill = table_header_fill
                cell.border = thin_border
            
            # Case type data
            case_types = cases_queryset.values('case_type').annotate(
                count=Count('case_type')
            ).order_by('-count')
            
            for case_type in case_types:
                current_row += 1
                display_name = dict(DisciplinaryCase.CASE_TYPE_CHOICES).get(
                    case_type['case_type'], 
                    case_type['case_type']
                )
                percentage = round((case_type['count'] / total_cases * 100), 1)
                
                ws[f'A{current_row}'] = display_name
                ws[f'B{current_row}'] = case_type['count']
                ws[f'C{current_row}'] = f"{percentage}%"
                
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{current_row}'].border = thin_border
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        filename = f"monthly_report_{calendar.month_name[month].lower()}_{year}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except ImportError:
        # Handle case where openpyxl is not installed
        messages.error(request, "Excel export requires openpyxl library. Please install it with: pip install openpyxl")
        return redirect('monthly_report')
    
    except Exception as e:
        # Handle any other errors during Excel generation
        messages.error(request, f"Error generating Excel report: {str(e)}")
        return redirect('monthly_report')
    
def student_profile_view(request):
    """Student profile view - displays complete student information"""
    if not request.session.get('is_student'):
        messages.error(request, 'Please log in to access your profile.')
        return redirect('student_login')
    
    try:
        # Get student from session
        student_id = request.session.get('student_id')
        student = get_object_or_404(Student, id=student_id)
        
        # Get student's disciplinary cases
        disciplinary_cases = DisciplinaryCase.objects.filter(
            student=student
        ).order_by('-incident_date')
        
        # Calculate statistics
        total_cases = disciplinary_cases.count()
        pending_cases = disciplinary_cases.filter(status='pending').count()
        under_review_cases = disciplinary_cases.filter(status='under_review').count()
        resolved_cases = disciplinary_cases.filter(status='resolved').count()
        
        # Remove total_points calculation since it's not needed
        
        # Get login history (optional, handle if StudentLogin doesn't exist)
        try:
            student_login = StudentLogin.objects.get(student=student)
            last_login = student_login.last_login
            login_count = student_login.login_count
        except:
            last_login = None
            login_count = 0
        
        context = {
            'student': student,
            'disciplinary_cases': disciplinary_cases,
            'total_cases': total_cases,
            'pending_cases': pending_cases,
            'under_review_cases': under_review_cases,
            'resolved_cases': resolved_cases,
            'last_login': last_login,
            'login_count': login_count,
        }
        
        return render(request, 'student_profile.html', context)
        
    except Exception as e:
        print(f"Profile error: {str(e)}")  # For debugging
        messages.error(request, 'Error loading profile. Please try again.')
        return redirect('student_login')  
        
    except Exception as e:
        print(f"Profile error: {str(e)}")  # For debugging
        messages.error(request, 'Error loading profile. Please try again.')
        return redirect('student_dashboard')

# Add this view function to your accounts/views.py

def student_cases_view(request):
    """Student cases view - displays all disciplinary cases for the logged-in student"""
    if not request.session.get('is_student'):
        messages.error(request, 'Please log in to access your cases.')
        return redirect('student_login')
    
    try:
        # Get student from session
        student_id = request.session.get('student_id')
        student = get_object_or_404(Student, id=student_id)
        
        # Get student's disciplinary cases
        disciplinary_cases = DisciplinaryCase.objects.filter(
            student=student
        ).order_by('-incident_date')
        
        context = {
            'student': student,
            'disciplinary_cases': disciplinary_cases,
        }
        
        return render(request, 'student_cases.html', context)
        
    except Exception as e:
        print(f"Cases error: {str(e)}")  # For debugging
        messages.error(request, 'Error loading cases. Please try again.')
        return redirect('student_profile')
# Add this view function to your accounts/views.py
# Add this view function to your accounts/views.py

